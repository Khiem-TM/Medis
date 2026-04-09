from __future__ import annotations

import io
import json
import logging
from datetime import date, datetime, timezone
from typing import List, Optional
from uuid import uuid4

import openpyxl
import openpyxl.styles
from fastapi import HTTPException, status
from openai import AsyncOpenAI
from sqlalchemy import and_, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.models.drug import Drug, DrugInteraction, InteractionSeverity
from app.models.health_profile import HealthProfile
from app.models.prescription import Prescription, PrescriptionStatus
from app.models.user import User
from app.schemas.recommendation import (
    DrugSuggestion,
    RecommendationRequest,
    RecommendationResponse,
)

logger = logging.getLogger(__name__)

_BLUE_FILL = openpyxl.styles.PatternFill("solid", fgColor="2563EB")
_WHITE_FONT = openpyxl.styles.Font(bold=True, color="FFFFFF")
_RED_FILL = openpyxl.styles.PatternFill("solid", fgColor="FEE2E2")

_SYSTEM_PROMPT = (
    "Bạn là AI dược sĩ tư vấn. Trả về JSON ONLY, không có text khác. "
    "Không kê đơn cho bệnh nguy hiểm. Luôn có cảnh báo tham khảo bác sĩ."
)

_USER_PROMPT_TEMPLATE = """Thông tin bệnh nhân:
- Tuổi: {age}, Giới tính: {gender}
- Triệu chứng hiện tại: {symptoms}
- Lịch sử bệnh: {diagnoses}
- Thuốc đang dùng: {current_drugs}

Hãy gợi ý 3-5 thuốc phù hợp nhất.
Trả về JSON ONLY, không có markdown, không có text khác.
Format:
{{
  "drugs": [
    {{
      "drug_name": "tên thuốc tiếng Việt hoặc generic name",
      "active_ingredient": "hoạt chất chính",
      "indication": "công dụng với triệu chứng này",
      "reference_dosage": "liều dùng tham khảo",
      "suitability_score": 0-100,
      "warnings": "cảnh báo quan trọng hoặc null"
    }}
  ],
  "general_advice": "lời khuyên chung cho tình trạng này",
  "see_doctor_if": "nên gặp bác sĩ nếu..."
}}"""


def _calc_age(dob: Optional[datetime]) -> Optional[int]:
    if not dob:
        return None
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


class RecommendationService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self.client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

    async def recommend(
        self,
        user_id: int,
        data: RecommendationRequest,
    ) -> RecommendationResponse:
        # ── Step 1: Collect context ──────────────────────────────────────
        user = await self.db.scalar(select(User).where(User.id == user_id))
        age = _calc_age(user.date_of_birth) if user else None
        gender = user.gender if user else "N/A"

        diagnoses = "Không có"
        if data.health_profile_ids:
            hp_rows = (
                await self.db.execute(
                    select(HealthProfile).where(
                        HealthProfile.id.in_(data.health_profile_ids),
                        HealthProfile.user_id == user_id,
                    )
                )
            ).scalars().all()
            # Ensure all requested profiles belong to user
            found_ids = {hp.id for hp in hp_rows}
            missing = [i for i in data.health_profile_ids if i not in found_ids]
            if missing:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Một số hồ sơ bệnh án không thuộc về tài khoản này",
                )
            if hp_rows:
                diagnoses = "; ".join(
                    f"{hp.diagnosis_name}" + (f" ({hp.symptoms})" if hp.symptoms else "")
                    for hp in hp_rows
                )

        current_drug_names = "Không có"
        current_drug_ids: List[str] = []
        if data.current_prescription_id:
            pres = await self.db.scalar(
                select(Prescription)
                .where(
                    Prescription.id == data.current_prescription_id,
                    Prescription.user_id == user_id,
                )
                .options(selectinload(Prescription.items))
            )
            if not pres:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Đơn thuốc không thuộc về tài khoản này",
                )
            current_drug_names = ", ".join(
                f"{item.drug_name} {item.dosage}" for item in pres.items
            ) or "Không có"
            current_drug_ids = [item.drug_id for item in pres.items if item.drug_id]

        # ── Step 2: Call OpenAI ──────────────────────────────────────────
        prompt = _USER_PROMPT_TEMPLATE.format(
            age=age or "N/A",
            gender=gender or "N/A",
            symptoms=data.symptoms,
            diagnoses=diagnoses,
            current_drugs=current_drug_names,
        )

        ai_response = await self.client.chat.completions.create(
            model=settings.OPENAI_MODEL,
            messages=[
                {"role": "system", "content": _SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ],
            max_tokens=1200,
            temperature=0.3,
        )

        raw = ai_response.choices[0].message.content.strip()
        # Strip possible markdown fences
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[-1]
            raw = raw.rsplit("```", 1)[0].strip()

        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="AI không thể tạo gợi ý, vui lòng thử lại",
            )

        # ── Step 3: Enrich with DB drug ids ─────────────────────────────
        suggestions: List[DrugSuggestion] = []
        for drug_data in parsed.get("drugs", []):
            drug_name = drug_data.get("drug_name", "")
            db_drug = await self.db.scalar(
                select(Drug)
                .where(Drug.name.ilike(f"%{drug_name}%"))
                .limit(1)
            )
            suggestions.append(
                DrugSuggestion(
                    drug_name=drug_name,
                    active_ingredient=drug_data.get("active_ingredient", ""),
                    indication=drug_data.get("indication", ""),
                    reference_dosage=drug_data.get("reference_dosage", ""),
                    suitability_score=int(drug_data.get("suitability_score", 0)),
                    warnings=drug_data.get("warnings") or None,
                    drug_id=db_drug.id if db_drug else None,
                )
            )

        # ── Step 4: Check interactions ───────────────────────────────────
        if current_drug_ids and suggestions:
            suggested_drug_ids = [s.drug_id for s in suggestions if s.drug_id]
            if suggested_drug_ids:
                conditions = []
                for cur_id in current_drug_ids:
                    for sug_id in suggested_drug_ids:
                        a, b = sorted([cur_id, sug_id])
                        conditions.append(
                            and_(
                                DrugInteraction.drug_id_1 == a,
                                DrugInteraction.drug_id_2 == b,
                            )
                        )
                if conditions:
                    interactions = (
                        await self.db.execute(
                            select(DrugInteraction).where(or_(*conditions))
                        )
                    ).scalars().all()

                    dangerous = {
                        i.drug_id_1 for i in interactions
                        if i.severity in (InteractionSeverity.major, InteractionSeverity.moderate)
                    } | {
                        i.drug_id_2 for i in interactions
                        if i.severity in (InteractionSeverity.major, InteractionSeverity.moderate)
                    }
                    for s in suggestions:
                        if s.drug_id and s.drug_id in dangerous:
                            s.has_interaction = True

        return RecommendationResponse(
            request_id=str(uuid4()),
            symptoms=data.symptoms,
            suggestions=suggestions,
            general_advice=parsed.get("general_advice", ""),
            see_doctor_if=parsed.get("see_doctor_if", ""),
            generated_at=datetime.now(timezone.utc),
        )

    async def export_xlsx(
        self,
        result: RecommendationResponse,
        user_id: int,
    ) -> bytes:
        user = await self.db.scalar(select(User).where(User.id == user_id))
        age = _calc_age(user.date_of_birth) if user else None
        full_name = user.full_name if user else "N/A"

        wb = openpyxl.Workbook()

        # ── Sheet 1: Thông tin ───────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Thông tin"
        info_rows = [
            ("Họ tên", full_name),
            ("Tuổi", str(age) if age else "N/A"),
            ("Triệu chứng", result.symptoms),
            ("Ngày tạo", result.generated_at.strftime("%d/%m/%Y %H:%M:%S")),
            ("Lời khuyên", result.general_advice),
            ("Nên gặp bác sĩ nếu", result.see_doctor_if),
        ]
        for label, value in info_rows:
            ws1.append([label, value])
        for col in ws1.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=15)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        # ── Sheet 2: Danh sách thuốc gợi ý ──────────────────────────────
        ws2 = wb.create_sheet("Danh sách thuốc gợi ý")
        headers = [
            "STT", "Tên thuốc", "Hoạt chất", "Công dụng",
            "Liều lượng tham khảo", "Mức phù hợp (%)",
            "Cảnh báo tương tác", "Cảnh báo khác",
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = _WHITE_FONT
            cell.fill = _BLUE_FILL

        for stt, s in enumerate(result.suggestions, 1):
            row_data = [
                stt,
                s.drug_name,
                s.active_ingredient,
                s.indication,
                s.reference_dosage,
                f"{s.suitability_score}%",
                "Có tương tác" if s.has_interaction else "",
                s.warnings or "",
            ]
            ws2.append(row_data)
            if s.has_interaction:
                for col_idx in range(1, len(headers) + 1):
                    ws2.cell(row=stt + 1, column=col_idx).fill = _RED_FILL

        for col in ws2.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=15)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
