from __future__ import annotations

import io
import json
import logging
from datetime import datetime
from math import ceil
from typing import Optional

import openpyxl
import openpyxl.styles
from fastapi import Request
from sqlalchemy import cast, delete as sql_delete, func, select, String
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.log import ActivityLog, LogLevel, SystemLog
from app.schemas.log import ActivityLogResponse, SystemLogResponse
from app.schemas.user import PaginatedResponse, PaginationMeta

logger = logging.getLogger(__name__)

_ACTION_MAP: dict[str, str] = {
    "LOGIN": "Đăng nhập",
    "LOGOUT": "Đăng xuất",
    "REGISTER": "Đăng ký tài khoản",
    "PROFILE_UPDATE": "Cập nhật hồ sơ",
    "PASSWORD_CHANGE": "Đổi mật khẩu",
    "DRUG_SEARCH": "Tra cứu thuốc",
    "INTERACTION_CHECK": "Kiểm tra tương tác thuốc",
    "PRESCRIPTION_CREATE": "Tạo đơn thuốc",
    "PRESCRIPTION_UPDATE": "Cập nhật đơn thuốc",
    "PRESCRIPTION_DELETE": "Xóa đơn thuốc",
    "HEALTH_PROFILE_CREATE": "Tạo hồ sơ bệnh án",
    "HEALTH_PROFILE_UPDATE": "Cập nhật hồ sơ bệnh án",
    "HEALTH_PROFILE_DELETE": "Xóa hồ sơ bệnh án",
    "CHATBOT_MESSAGE": "Tư vấn chatbot",
}

_BLUE_FILL = openpyxl.styles.PatternFill("solid", fgColor="2563EB")
_WHITE_FONT = openpyxl.styles.Font(bold=True, color="FFFFFF")


# ══════════════════════════════════════════════════════════════════════════════
#  ActivityLogService
# ══════════════════════════════════════════════════════════════════════════════

class ActivityLogService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def log(
        self,
        action: str,
        user_id: Optional[int] = None,
        entity_type: Optional[str] = None,
        entity_id: Optional[str] = None,
        detail: Optional[dict] = None,
        request: Optional[Request] = None,
    ) -> None:
        try:
            ip = request.client.host if request and request.client else None
            ua = request.headers.get("user-agent", "") if request else None
            entry = ActivityLog(
                user_id=user_id,
                action=action,
                entity_type=entity_type,
                entity_id=entity_id,
                detail=detail,
                ip_address=ip,
                user_agent=ua,
            )
            self.db.add(entry)
            await self.db.flush()
        except Exception as exc:
            logger.error("ActivityLogService.log failed: %s", exc)

    async def get_user_history(
        self,
        user_id: int,
        page: int,
        size: int,
        action: Optional[str] = None,
        keyword: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> PaginatedResponse:
        stmt = select(ActivityLog).where(ActivityLog.user_id == user_id)

        if action:
            stmt = stmt.where(ActivityLog.action == action)
        if keyword:
            stmt = stmt.where(
                cast(ActivityLog.detail, String).ilike(f"%{keyword}%")
            )
        if date_from:
            stmt = stmt.where(ActivityLog.created_at >= date_from)
        if date_to:
            stmt = stmt.where(ActivityLog.created_at <= date_to)

        total = await self.db.scalar(
            select(func.count()).select_from(stmt.subquery())
        )
        stmt = (
            stmt.order_by(ActivityLog.created_at.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
        rows = (await self.db.execute(stmt)).scalars().all()

        return PaginatedResponse(
            items=[ActivityLogResponse.model_validate(r) for r in rows],
            meta=PaginationMeta(
                total=total,
                page=page,
                size=size,
                total_pages=max(1, ceil(total / size)),
            ),
        )

    async def get_all_history(
        self,
        page: int,
        size: int,
        user_id: Optional[int] = None,
        action: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> PaginatedResponse:
        """Admin: query across all users."""
        stmt = select(ActivityLog)
        if user_id is not None:
            stmt = stmt.where(ActivityLog.user_id == user_id)
        if action:
            stmt = stmt.where(ActivityLog.action == action)
        if date_from:
            stmt = stmt.where(ActivityLog.created_at >= date_from)
        if date_to:
            stmt = stmt.where(ActivityLog.created_at <= date_to)

        total = await self.db.scalar(
            select(func.count()).select_from(stmt.subquery())
        )
        stmt = (
            stmt.order_by(ActivityLog.created_at.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
        rows = (await self.db.execute(stmt)).scalars().all()

        return PaginatedResponse(
            items=[ActivityLogResponse.model_validate(r) for r in rows],
            meta=PaginationMeta(
                total=total,
                page=page,
                size=size,
                total_pages=max(1, ceil(total / size)),
            ),
        )

    async def delete_one(self, user_id: int, log_id: int) -> None:
        from fastapi import HTTPException, status
        result = await self.db.execute(
            select(ActivityLog).where(
                ActivityLog.id == log_id, ActivityLog.user_id == user_id
            )
        )
        entry = result.scalar_one_or_none()
        if not entry:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Không tìm thấy bản ghi",
            )
        await self.db.delete(entry)
        await self.db.flush()

    async def delete_all(self, user_id: int) -> int:
        result = await self.db.execute(
            sql_delete(ActivityLog).where(ActivityLog.user_id == user_id)
        )
        return result.rowcount

    async def export_xlsx(self, user_id: Optional[int] = None) -> bytes:
        stmt = select(ActivityLog)
        if user_id is not None:
            stmt = stmt.where(ActivityLog.user_id == user_id)
        stmt = stmt.order_by(ActivityLog.created_at.desc())
        rows = (await self.db.execute(stmt)).scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch sử hoạt động"

        headers = ["STT", "Thời gian", "Loại hoạt động", "Đối tượng", "Chi tiết"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _WHITE_FONT
            cell.fill = _BLUE_FILL

        for stt, row in enumerate(rows, 1):
            entity = ""
            if row.entity_type and row.entity_id:
                entity = f"{row.entity_type} #{row.entity_id}"
            detail_str = ""
            if row.detail:
                detail_str = json.dumps(row.detail, ensure_ascii=False)
            ws.append([
                stt,
                row.created_at.strftime("%d/%m/%Y %H:%M:%S") if row.created_at else "",
                _ACTION_MAP.get(row.action, row.action),
                entity,
                detail_str,
            ])

        _autofit(ws, min_width=15, max_width=50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  SystemLogService
# ══════════════════════════════════════════════════════════════════════════════

_LEVEL_COLORS: dict[str, str] = {
    "ERROR": "FEE2E2",
    "CRITICAL": "FEE2E2",
    "WARNING": "FEF9C3",
    "INFO": "DBEAFE",
    "DEBUG": "DBEAFE",
}


class SystemLogService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def log(
        self,
        level: str,
        source: str,
        message: str,
        detail: Optional[dict] = None,
    ) -> None:
        try:
            entry = SystemLog(
                level=LogLevel(level),
                source=source,
                message=message,
                detail=detail,
            )
            self.db.add(entry)
            await self.db.flush()
        except Exception as exc:
            logger.error("SystemLogService.log failed: %s", exc)

    async def get_list(
        self,
        page: int,
        size: int,
        level: Optional[str] = None,
        source: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> PaginatedResponse:
        stmt = select(SystemLog)
        if level:
            stmt = stmt.where(SystemLog.level == LogLevel(level))
        if source:
            stmt = stmt.where(SystemLog.source.ilike(f"%{source}%"))
        if date_from:
            stmt = stmt.where(SystemLog.created_at >= date_from)
        if date_to:
            stmt = stmt.where(SystemLog.created_at <= date_to)

        total = await self.db.scalar(
            select(func.count()).select_from(stmt.subquery())
        )
        stmt = (
            stmt.order_by(SystemLog.created_at.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
        rows = (await self.db.execute(stmt)).scalars().all()

        return PaginatedResponse(
            items=[SystemLogResponse.model_validate(r) for r in rows],
            meta=PaginationMeta(
                total=total,
                page=page,
                size=size,
                total_pages=max(1, ceil(total / size)),
            ),
        )

    async def export_xlsx(self) -> bytes:
        rows = (
            await self.db.execute(
                select(SystemLog).order_by(SystemLog.created_at.desc())
            )
        ).scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "System Logs"

        headers = ["STT", "Thời gian", "Level", "Source", "Message", "Chi tiết"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _WHITE_FONT
            cell.fill = _BLUE_FILL

        for stt, row in enumerate(rows, 1):
            detail_str = json.dumps(row.detail, ensure_ascii=False) if row.detail else ""
            data_row = ws.max_row + 1
            ws.append([
                stt,
                row.created_at.strftime("%d/%m/%Y %H:%M:%S") if row.created_at else "",
                row.level.value if hasattr(row.level, "value") else str(row.level),
                row.source,
                row.message,
                detail_str,
            ])
            color = _LEVEL_COLORS.get(
                row.level.value if hasattr(row.level, "value") else str(row.level),
                "DBEAFE",
            )
            fill = openpyxl.styles.PatternFill("solid", fgColor=color)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=data_row, column=col_idx).fill = fill

        _autofit(ws, min_width=15, max_width=50)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ── Helpers ────────────────────────────────────────────────────────────────

def _autofit(ws, min_width: int = 15, max_width: int = 50) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value or "")) for cell in col), default=min_width
        )
        ws.column_dimensions[col[0].column_letter].width = min(
            max(max_len + 2, min_width), max_width
        )
